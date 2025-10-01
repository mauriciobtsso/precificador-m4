from flask import (
    render_template, request, redirect, url_for, flash, send_from_directory
)
from flask_login import login_required
from app import db
from app.models import Produto
from openpyxl import load_workbook
from io import TextIOWrapper
import csv
import os

# utilitários que você já usava em routes.py
from app.utils.numeros import to_float, to_number
from app.utils.importar import _headers_lower, _row_as_dict, _get
from app.utils.db_helpers import get_or_404

from . import produtos_bp


###########################
# --- Produtos ---
###########################

@produtos_bp.route("/", methods=["GET"])
@login_required
def produtos():
    termo = request.args.get("termo", "").strip()
    lucro = request.args.get("lucro")
    preco_min = request.args.get("preco_min")
    preco_max = request.args.get("preco_max")

    query = db.session.query(Produto)

    if termo:
        like = f"%{termo}%"
        query = query.filter(
            (Produto.nome.ilike(like)) | (Produto.sku.ilike(like))
        )

    if lucro == "positivo":
        query = query.filter(Produto.lucro_liquido_real >= 0)
    elif lucro == "negativo":
        query = query.filter(Produto.lucro_liquido_real < 0)

    if preco_min:
        try:
            query = query.filter(Produto.preco_a_vista >= float(preco_min))
        except:
            pass
    if preco_max:
        try:
            query = query.filter(Produto.preco_a_vista <= float(preco_max))
        except:
            pass

    produtos = query.all()
    return render_template("produtos.html", produtos=produtos)


@produtos_bp.route("/novo", methods=["GET", "POST"])
@produtos_bp.route("/editar/<int:produto_id>", methods=["GET", "POST"])
@login_required
def gerenciar_produto(produto_id=None):
    produto = db.session.get(Produto, produto_id) if produto_id else None

    if request.method == "POST":
        if not produto:
            produto = Produto()
            db.session.add(produto)

        produto.sku = request.form.get("sku", "").upper()
        produto.nome = request.form["nome"]

        produto.preco_fornecedor = to_float(request.form.get("preco_fornecedor"))
        produto.desconto_fornecedor = to_float(request.form.get("desconto_fornecedor"))

        produto.margem = to_float(request.form.get("margem"))
        produto.lucro_alvo = (to_float(request.form.get("lucro_alvo")) or None)
        produto.preco_final = (to_float(request.form.get("preco_final")) or None)

        # 🔹 Novo campo frete
        produto.frete = to_float(request.form.get("frete"))

        produto.ipi = to_float(request.form.get("ipi"))
        produto.ipi_tipo = request.form.get("ipi_tipo", "%")
        produto.difal = to_float(request.form.get("difal"))
        produto.imposto_venda = to_float(request.form.get("imposto_venda"))

        # 🔹 Calcula preços com frete incluído
        produto.calcular_precos()

        db.session.commit()
        flash("Produto salvo com sucesso!", "success")
        return redirect(url_for("produtos.produtos"))

    return render_template("produto_form.html", produto=produto)


@produtos_bp.route("/excluir/<int:produto_id>")
@login_required
def excluir_produto(produto_id):
    produto = get_or_404(Produto, produto_id)
    db.session.delete(produto)
    db.session.commit()
    flash("Produto excluído com sucesso!", "success")
    return redirect(url_for("produtos.produtos"))


# --- Importação de Produtos (sem pandas) ---
@produtos_bp.route("/importar", methods=["GET", "POST"])
@login_required
def importar_produtos():
    if request.method == "POST":
        file = request.files.get("arquivo")
        if not file or file.filename == "":
            flash("Nenhum arquivo selecionado.", "warning")
            return redirect(url_for("produtos.importar_produtos"))

        filename = file.filename.lower()

        try:
            criados, atualizados = 0, 0

            if filename.endswith(".xlsx"):
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                headers = _headers_lower(ws)

                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = _row_as_dict(headers, row)

                    sku = (str(_get(data, "sku") or "").strip().upper())
                    if not sku:
                        continue

                    produto = db.session.query(Produto).filter_by(sku=sku).first()
                    if not produto:
                        produto = Produto(sku=sku)
                        db.session.add(produto)
                        criados += 1
                    else:
                        atualizados += 1

                    produto.nome = _get(data, "nome", default=produto.nome)
                    produto.preco_fornecedor = to_number(_get(data, "preco_fornecedor"))
                    produto.desconto_fornecedor = to_number(_get(data, "desconto_fornecedor"))
                    produto.margem = to_number(_get(data, "margem"))
                    produto.lucro_alvo = to_number(_get(data, "lucro_alvo")) or None
                    produto.preco_final = to_number(_get(data, "preco_final")) or None

                    produto.ipi = to_number(_get(data, "ipi"))
                    produto.ipi_tipo = _get(data, "ipi_tipo", default="%") or "%"
                    produto.difal = to_number(_get(data, "difal"))
                    produto.imposto_venda = to_number(_get(data, "imposto_venda"))

                    produto.calcular_precos()

            elif filename.endswith(".csv"):
                # Detectar encoding/delimitador e ler linhas
                text = TextIOWrapper(file.stream, encoding="utf-8-sig", newline="")
                try:
                    sniffer = csv.Sniffer()
                    sample = text.read(2048)
                    text.seek(0)
                    dialect = sniffer.sniff(sample)
                except Exception:
                    text.seek(0)
                    dialect = csv.excel

                reader = csv.DictReader(text, dialect=dialect)
                # normalizar headers para lower
                field_map = {h: h.strip().lower() for h in reader.fieldnames or []}

                for row in reader:
                    data = {field_map.get(k, k).lower(): v for k, v in row.items()}

                    sku = (str(data.get("sku") or "").strip().upper())
                    if not sku:
                        continue

                    produto = db.session.query(Produto).filter_by(sku=sku).first()
                    if not produto:
                        produto = Produto(sku=sku)
                        db.session.add(produto)
                        criados += 1
                    else:
                        atualizados += 1

                    produto.nome = data.get("nome", produto.nome)
                    produto.preco_fornecedor = to_number(data.get("preco_fornecedor"))
                    produto.desconto_fornecedor = to_number(data.get("desconto_fornecedor"))
                    produto.margem = to_number(data.get("margem"))
                    produto.lucro_alvo = to_number(data.get("lucro_alvo")) or None
                    produto.preco_final = to_number(data.get("preco_final")) or None

                    produto.ipi = to_number(data.get("ipi"))
                    produto.ipi_tipo = data.get("ipi_tipo") or "%"
                    produto.difal = to_number(data.get("difal"))
                    produto.imposto_venda = to_number(data.get("imposto_venda"))

                    produto.calcular_precos()
            else:
                flash("Formato de arquivo não suportado. Use .csv ou .xlsx", "danger")
                return redirect(url_for("produtos.importar_produtos"))

            db.session.commit()
            flash(f"Importação concluída! {criados} criados, {atualizados} atualizados.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao importar: {e}", "danger")

        return redirect(url_for("produtos.produtos"))

    return render_template("produtos_importar.html")


@produtos_bp.route("/exemplo-csv")
@login_required
def exemplo_csv():
    pasta = os.path.join(os.path.dirname(__file__), "..", "static")
    return send_from_directory(pasta, "exemplo_produtos.csv", as_attachment=True)
