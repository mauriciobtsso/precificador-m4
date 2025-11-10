# ================================================================
# app/services/importacao.py
# Revisado — compatível com app/vendas/models.py e relatórios TDVendas
# ================================================================

from openpyxl import load_workbook
from app.extensions import db
from app.clientes.models import Cliente
from app.vendas.models import Venda, ItemVenda  # <-- atualizado
from app.utils.excel_helpers import _headers_lower, _row_as_dict, _get, _as_bool
from app.utils.number_helpers import to_float
from app.utils.date_helpers import parse_data


# =====================================================
# IMPORTAÇÃO DE CLIENTES
# =====================================================
def importar_clientes(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    headers = _headers_lower(ws)

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = _row_as_dict(headers, row)

        nome = _get(data, "nome razão social", "nome")
        if not nome:
            continue

        doc = str(_get(data, "documento (cpf / cnpj)", "documento") or "").strip()
        if not doc:
            doc = None

        cliente = Cliente.query.filter_by(documento=doc).first() if doc else None
        if not cliente:
            cliente = Cliente(nome=nome, documento=doc)
            db.session.add(cliente)

        cliente.nome = nome
        cliente.razao_social = _get(data, "razão social", "razao social", default="")
        cliente.sexo = _get(data, "sexo", default="")
        cliente.profissao = _get(data, "profissão", "profissao", default="")
        cliente.rg = _get(data, "rg", default="")
        cliente.rg_emissor = _get(data, "rg emissor", default="")
        cliente.email = _get(data, "e-mail", "email", default="")
        cliente.telefone = _get(data, "telefone", default="")
        cliente.celular = _get(data, "celular", default="")
        cliente.endereco = _get(data, "endereço", "endereco", default="")
        cliente.numero = str(_get(data, "número", "numero", default="") or "")
        cliente.complemento = _get(data, "complemento", default="")
        cliente.bairro = _get(data, "bairro", default="")
        cliente.cidade = _get(data, "cidade", default="")
        cliente.estado = _get(data, "estado", default="")
        cliente.cep = _get(data, "cep", default="")
        cliente.cr = _get(data, "cr", default="")
        cliente.cr_emissor = _get(data, "cr emissor", default="")
        cliente.sigma = _get(data, "sigma", default="")
        cliente.sinarm = _get(data, "sinarm", default="")
        cliente.cac = _as_bool(_get(data, "cac"))
        cliente.filiado = _as_bool(_get(data, "filiado"))
        cliente.policial = _as_bool(_get(data, "policial"))
        cliente.bombeiro = _as_bool(_get(data, "bombeiro"))
        cliente.militar = _as_bool(_get(data, "militar"))
        cliente.iat = _as_bool(_get(data, "iat"))
        cliente.psicologo = _as_bool(_get(data, "psicologo"))

    db.session.commit()


# =====================================================
# IMPORTAÇÃO DE VENDAS (Relatórios TDVendas)
# =====================================================
def importar_vendas(file_storage):
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    headers = _headers_lower(ws)

    current_venda = None
    ultima_chave_venda = None  # identifica unicamente cada venda

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = _row_as_dict(headers, row)

        consumidor = _get(data, "consumidor")
        documento = str(_get(data, "documento") or "").strip() or None
        abertura = parse_data(_get(data, "abertura"))
        nf_numero = str(_get(data, "nf - nº", "nf-nº", "nf nº", default="") or "")

        # Gera uma "chave lógica" única para cada venda
        chave_atual = f"{consumidor}|{abertura}|{nf_numero}"

        # =============== NOVA VENDA ===============
        if consumidor and (chave_atual != ultima_chave_venda):
            # --- Localiza ou cria cliente ---
            if documento:
                cliente = Cliente.query.filter_by(documento=documento).first()
                if not cliente:
                    cliente = Cliente(nome=consumidor or "", documento=documento)
                    db.session.add(cliente)
                    db.session.flush()
            else:
                cliente = Cliente.query.filter_by(documento=None, nome="Consumidor não identificado").first()
                if not cliente:
                    cliente = Cliente(nome="Consumidor não identificado", documento=None)
                    db.session.add(cliente)
                    db.session.flush()

            # --- Cria nova venda ---
            current_venda = Venda(
                cliente_id=cliente.id,
                vendedor=_get(data, "vendedor"),
                caixa=_get(data, "caixa"),
                status=_get(data, "status"),
                status_financeiro=_get(data, "status financeiro"),
                data_abertura=abertura,
                data_fechamento=parse_data(_get(data, "fechamento")),
                data_quitacao=parse_data(_get(data, "quitação", "quitacao")),
                data_cancelamento=parse_data(_get(data, "cancelamento")),
                valor_total=0.0,
                desconto_valor=to_float(_get(data, "descontos (r$)")),
                desconto_percentual=to_float(_get(data, "descontos (%)")),
                valor_recebido=to_float(_get(data, "valor recebido")),
                valor_faltante=to_float(_get(data, "valor faltante")),
                crediario=_as_bool(_get(data, "crediário")),
                parcelas_qtd=int(to_float(_get(data, "parcelas - qtd", default=0)) or 0),
                parcelas_primeiro_vencimento=parse_data(_get(data, "parcelas - primeiro vencimento")),
                parcelas_ultimo_vencimento=parse_data(_get(data, "parcelas - ultimo vencimento")),
                nf_data=parse_data(_get(data, "nf - data")),
                nf_numero=nf_numero,
                nf_valor=to_float(_get(data, "nf - valor", "nf valor")),
                teve_devolucao=_as_bool(_get(data, "teve devoluções", "teve devolucoes")),
                cliente_nome=consumidor,
                documento_cliente=documento,
                tipo_pessoa=_get(data, "pessoa"),
            )
            db.session.add(current_venda)
            db.session.flush()

            ultima_chave_venda = chave_atual

        # =============== ITEM DA VENDA (PRIMEIRO OU ADICIONAL) ===============
        if current_venda and _get(data, "produto"):
            qtd = int(to_float(_get(data, "itens - qtd", "qtd", default=1)) or 1)
            valor = to_float(_get(data, "valor"))
            subtotal = valor * qtd

            item = ItemVenda(
                venda_id=current_venda.id,
                produto_nome=_get(data, "produto", default=""),
                categoria=_get(data, "tipo do produto", "categoria", default=""),
                quantidade=qtd,
                valor_unitario=valor,
                valor_total=subtotal,
            )
            db.session.add(item)

            # Atualiza valor total acumulado da venda
            current_venda.valor_total = (current_venda.valor_total or 0) + subtotal

    db.session.commit()
