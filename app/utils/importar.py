# ======================
# IMPORTAÇÃO DE PLANILHAS DE PRODUTOS
# ======================

import csv
import io
from decimal import Decimal
from openpyxl import load_workbook
from app import db
from app.produtos.models import Produto


def _headers_lower(ws):
    """Normaliza os cabeçalhos da planilha para minúsculo."""
    headers = []
    for cell in ws[1]:
        if cell.value:
            headers.append(str(cell.value).strip().lower())
        else:
            headers.append("")
    return headers


def _row_as_dict(headers, row):
    """Converte uma linha da planilha em dict usando os headers."""
    data = {}
    for header, value in zip(headers, row):
        data[header] = value
    return data


def _get(data: dict, key: str, default=None):
    """Acessa uma chave em dict com segurança, retornando default se não existir."""
    try:
        return data.get(key, default)
    except Exception:
        return default


# ======================
# FUNÇÃO PRINCIPAL
# ======================
def importar_planilha_produtos(arquivo):
    """
    Lê um arquivo CSV ou XLSX e importa produtos.
    Atualiza se já existir (mesmo código), senão cria novo.
    """
    nome = arquivo.filename.lower()
    produtos_importados = []

    if nome.endswith(".csv"):
        # Leitura CSV
        stream = io.StringIO(arquivo.stream.read().decode("utf-8-sig"))
        reader = csv.DictReader(stream)
        linhas = list(reader)
    elif nome.endswith(".xlsx"):
        # Leitura XLSX
        wb = load_workbook(arquivo, data_only=True)
        ws = wb.active
        headers = _headers_lower(ws)
        linhas = [_row_as_dict(headers, [cell.value for cell in row]) for row in ws.iter_rows(min_row=2)]
    else:
        raise ValueError("Formato de arquivo não suportado.")

    for linha in linhas:
        codigo = str(_get(linha, "sku") or _get(linha, "codigo") or "").strip().upper()
        if not codigo:
            continue

        produto = Produto.query.filter_by(codigo=codigo).first()
        if not produto:
            produto = Produto(codigo=codigo)

        produto.nome = str(_get(linha, "nome") or "").strip()
        produto.preco_fornecedor = _to_decimal(_get(linha, "preco_fornecedor"))
        produto.desconto_fornecedor = _to_decimal(_get(linha, "desconto_fornecedor"))
        produto.margem = _to_decimal(_get(linha, "margem"))
        produto.ipi = _to_decimal(_get(linha, "ipi"))
        produto.ipi_tipo = str(_get(linha, "ipi_tipo") or "%").strip()
        produto.difal = _to_decimal(_get(linha, "difal"))
        produto.imposto_venda = _to_decimal(_get(linha, "imposto_venda"))
        produto.frete = _to_decimal(_get(linha, "frete"))

        # Calcula preços e salva
        produto.calcular_precos()
        db.session.add(produto)
        produtos_importados.append(produto)

    db.session.commit()
    return produtos_importados


def _to_decimal(valor):
    """Converte valores numéricos em Decimal com segurança."""
    if valor is None or valor == "":
        return Decimal(0)
    try:
        if isinstance(valor, (float, int, Decimal)):
            return Decimal(valor)
        return Decimal(str(valor).replace(",", ".").strip())
    except Exception:
        return Decimal(0)
