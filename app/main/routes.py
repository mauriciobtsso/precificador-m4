from flask import render_template, request, redirect, url_for, flash, send_from_directory, current_app, send_file, jsonify
from flask_login import login_user, logout_user, login_required
from app.extensions import db
from app.models import Produto, Taxa, User, Configuracao, Venda, ItemVenda, PedidoCompra, ItemPedido
from app.clientes.models import Cliente
from app.main import main
from sqlalchemy import text, func, extract
from openpyxl import load_workbook
from io import TextIOWrapper
import csv
import os
from datetime import datetime, timedelta
from app.utils.gerar_pedidos import gerar_pedido_m4
import app.utils.parcelamento as parc
from app.utils.whatsapp import gerar_mensagem_whatsapp


# =====================================================
# Helpers
# =====================================================
def get_config(chave, default=None):
    conf = Configuracao.query.filter_by(chave=chave).first()
    return conf.valor if conf else default

def to_float(value, default=0.0):
    try:
        if value is None or str(value).strip() == "":
            return default
        return float(str(value).replace(",", "."))
    except (TypeError, ValueError):
        return default

def br_money(v: float) -> str:
    s = f"{float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"

def montar_parcelas(valor_base, taxas, modo="coeficiente_total"):
    resultado = []
    base = float(valor_base or 0)

    for taxa in taxas:
        n = int(taxa.numero_parcelas or 1)
        j = max(to_float(taxa.juros), 0.0) / 100.0
        if n <= 0:
            continue

        if n == 1 and modo == "coeficiente_total":
            coef = max(1.0 - j, 1e-9)
            total = base / coef
            parcela = total
        else:
            if modo == "juros_mensal":
                if j > 0:
                    parcela = base * (j / (1 - (1 + j) ** (-n)))
                else:
                    parcela = base / n
                total = parcela * n
            else:
                coef = max(1.0 - j, 1e-9)
                total = base / coef
                parcela = total / n

        resultado.append({
            "parcelas": n,
            "parcela": parcela,
            "total": total,
            "diferenca": total - base,
            "rotulo": f"{n}x",
        })

    return resultado

# =========================
# Função compor_whatsapp
# =========================
def compor_whatsapp(produto=None, valor_base=0.0, linhas=None):
    base = float(valor_base or 0)
    linhas = linhas or []

    prefixo = get_config("whatsapp_prefixo", "")

    cab = []
    if prefixo:
        cab.append(prefixo)

    if produto:
        cab.append(f"🔫 {produto.nome}")
        cab.append(f"🔖 SKU: {produto.sku}")
        cab.append(f"💰 À vista: {br_money(base)}")
    else:
        cab.append("💳 Simulação de Parcelamento")
        cab.append(f"💰 À vista: {br_money(base)}")

    corpo = []

    # ✅ PIX sempre fixo
    corpo.append(f"PIX {br_money(base)}")

    # ✅ percorre as linhas das taxas, mas ignora Pix duplicado
    for r in linhas:
        rotulo = r["rotulo"]

        if rotulo.lower() == "pix":
            continue  # já adicionamos acima

        if rotulo.lower() == "débito":
            # Débito mostra só o total (sem repetir =)
            corpo.append(f"Débito {br_money(r['total'])}")
        else:
            corpo.append(f"{rotulo} {br_money(r['parcela'])} = {br_money(r['total'])}")

    txt = "\n".join(cab) + "\n\n" + "💳 Opções de Parcelamento:\n" + "\n".join(corpo)
    txt += "\n\n⚠️ Os valores poderão sofrer alterações sem aviso prévio."
    return txt

def to_number(x):
    if isinstance(x, (int, float)):
        return float(x)
    if x is None:
        return 0.0
    s = str(x).strip()
    if not s:
        return 0.0
    s = s.replace("R$", "").replace(" ", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except:
        return 0.0

def parse_data(value):
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value
    if isinstance(value, (int, float)):
        try:
            base = datetime(1899, 12, 30)
            return base + timedelta(days=float(value))
        except Exception:
            pass
    for fmt in ("%d/%m/%Y %H:%M", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(str(value), fmt)
        except Exception:
            continue
    return None

def _headers_lower(ws):
    return [str(c.value).strip().lower() if c.value is not None else "" for c in ws[1]]

def _row_as_dict(headers_lower, row_values):
    return {h: v for h, v in zip(headers_lower, row_values) if h}

def _get(d, *names, default=None):
    for name in names:
        key = str(name).strip().lower()
        if key in d and d[key] not in (None, ""):
            return d[key]
    return default

def _as_bool(val):
    if val is None:
        return False
    s = str(val).strip().lower()
    return s in ("1", "sim", "true", "verdadeiro", "yes", "y")

# =====================================================
# Rotas principais
# =====================================================
@main.route("/")
def index():
    return redirect(url_for("main.dashboard"))

# --- Login ---
@main.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        # Busca o usuário pelo username
        user = User.query.filter_by(username=username).first()

        # Valida senha usando check_password
        if user and user.check_password(password):
            login_user(user)
            flash("Login realizado com sucesso!", "success")
            # Redireciona para a página que o usuário tentou acessar ou dashboard
            next_page = request.args.get("next")
            return redirect(next_page or url_for("main.dashboard"))
        else:
            flash("Usuário ou senha inválidos", "danger")

    return render_template("login.html")

# --- Logout ---
@main.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logout realizado com sucesso.", "info")
    return redirect(url_for("main.login"))
# --- Dashboard ---
@main.route("/dashboard")
@login_required
def dashboard():
    produtos = Produto.query.all()
    hoje = datetime.today()

    # Total de vendas no mês atual
    total_vendas_mes = (
        db.session.query(func.sum(Venda.valor_total))
        .filter(extract("year", Venda.data_abertura) == hoje.year)
        .filter(extract("month", Venda.data_abertura) == hoje.month)
        .scalar()
        or 0
    )

    # Top 5 clientes por valor de compras
    top_clientes = (
        db.session.query(
            Cliente.nome,
            func.sum(Venda.valor_total).label("total")
        )
        .join(Venda, Cliente.id == Venda.cliente_id)
        .group_by(Cliente.id)
        .order_by(func.sum(Venda.valor_total).desc())
        .limit(5)
        .all()
    )

    # Produto mais vendido (quantidade)
    produto_mais_vendido = (
        db.session.query(
            ItemVenda.produto_nome,
            func.sum(ItemVenda.quantidade).label("qtd")
        )
        .group_by(ItemVenda.produto_nome)
        .order_by(func.sum(ItemVenda.quantidade).desc())
        .first()
    )

    # Ticket médio
    ticket_medio = (
        db.session.query(func.sum(Venda.valor_total) / func.count(Venda.id))
        .scalar()
        or 0
    )

    # Vendas nos últimos 6 meses
    vendas_por_mes = (
        db.session.query(
            extract("month", Venda.data_abertura).label("mes"),
            func.sum(Venda.valor_total).label("total")
        )
        .filter(Venda.data_abertura >= hoje - timedelta(days=180))
        .group_by(extract("month", Venda.data_abertura))
        .order_by(extract("month", Venda.data_abertura))
        .all()
    )

    mapa_meses = [
        "Jan", "Fev", "Mar", "Abr", "Mai", "Jun",
        "Jul", "Ago", "Set", "Out", "Nov", "Dez"
    ]

    meses_nomes = [mapa_meses[int(m) - 1] for m, _ in vendas_por_mes]
    totais = [float(total) for _, total in vendas_por_mes]

    return render_template(
        "dashboard.html",
        produtos=produtos,
        total_vendas_mes=total_vendas_mes,
        top_clientes=top_clientes,
        produto_mais_vendido=produto_mais_vendido,
        ticket_medio=ticket_medio,
        meses=meses_nomes,
        totais=totais,
    )


# --- Produtos ---
@main.route("/produtos")
@login_required
def produtos():
    termo = request.args.get("termo", "").strip()
    lucro = request.args.get("lucro")
    preco_min = request.args.get("preco_min")
    preco_max = request.args.get("preco_max")

    query = Produto.query

    if termo:
        like = f"%{termo}%"
        query = query.filter(
            (Produto.nome.ilike(like)) | (Produto.sku.ilike(like))
        )

    if lucro == "positivo":
        query = query.filter(Produto.lucro_liquido_real >= 0)
    elif lucro == "negativo":
        query = query.filter(Produto.lucro_liquido_real < 0)

    if preco_min:
        try:
            query = query.filter(Produto.preco_a_vista >= float(preco_min))
        except:
            pass
    if preco_max:
        try:
            query = query.filter(Produto.preco_a_vista <= float(preco_max))
        except:
            pass

    produtos = query.all()
    return render_template("produtos.html", produtos=produtos)

@main.route("/produto/novo", methods=["GET", "POST"])
@main.route("/produto/editar/<int:produto_id>", methods=["GET", "POST"])
@login_required
def gerenciar_produto(produto_id=None):
    produto = Produto.query.get(produto_id) if produto_id else None
    if request.method == "POST":
        if not produto:
            produto = Produto()
            db.session.add(produto)

        produto.sku = request.form.get("sku", "").upper()
        produto.nome = request.form["nome"]

        produto.preco_fornecedor = to_float(request.form.get("preco_fornecedor"))
        produto.desconto_fornecedor = to_float(request.form.get("desconto_fornecedor"))

        produto.margem = to_float(request.form.get("margem"))
        produto.lucro_alvo = (to_float(request.form.get("lucro_alvo")) or None)
        produto.preco_final = (to_float(request.form.get("preco_final")) or None)

        # 🔹 Novo campo frete
        produto.frete = to_float(request.form.get("frete"))

        produto.ipi = to_float(request.form.get("ipi"))
        produto.ipi_tipo = request.form.get("ipi_tipo", "%")
        produto.difal = to_float(request.form.get("difal"))
        produto.imposto_venda = to_float(request.form.get("imposto_venda"))

        # 🔹 Calcula preços com frete incluído
        produto.calcular_precos()

        db.session.commit()
        flash("Produto salvo com sucesso!", "success")
        return redirect(url_for("main.produtos"))

    return render_template("produto_form.html", produto=produto)

@main.route("/produto/excluir/<int:produto_id>")
@login_required
def excluir_produto(produto_id):
    produto = Produto.query.get_or_404(produto_id)
    db.session.delete(produto)
    db.session.commit()
    flash("Produto excluído com sucesso!", "success")
    return redirect(url_for("main.produtos"))

# --- Importação de Produtos (sem pandas) ---
@main.route("/produtos/importar", methods=["GET", "POST"])
@login_required
def importar_produtos():
    if request.method == "POST":
        file = request.files.get("arquivo")
        if not file or file.filename == "":
            flash("Nenhum arquivo selecionado.", "warning")
            return redirect(url_for("main.importar_produtos"))

        filename = file.filename.lower()

        try:
            criados, atualizados = 0, 0

            if filename.endswith(".xlsx"):
                wb = load_workbook(file, data_only=True)
                ws = wb.active
                headers = _headers_lower(ws)

                for row in ws.iter_rows(min_row=2, values_only=True):
                    data = _row_as_dict(headers, row)

                    sku = (str(_get(data, "sku") or "").strip().upper())
                    if not sku:
                        continue

                    produto = Produto.query.filter_by(sku=sku).first()
                    if not produto:
                        produto = Produto(sku=sku)
                        db.session.add(produto)
                        criados += 1
                    else:
                        atualizados += 1

                    produto.nome = _get(data, "nome", default=produto.nome)
                    produto.preco_fornecedor = to_number(_get(data, "preco_fornecedor"))
                    produto.desconto_fornecedor = to_number(_get(data, "desconto_fornecedor"))
                    produto.margem = to_number(_get(data, "margem"))
                    produto.lucro_alvo = to_number(_get(data, "lucro_alvo")) or None
                    produto.preco_final = to_number(_get(data, "preco_final")) or None

                    produto.ipi = to_number(_get(data, "ipi"))
                    produto.ipi_tipo = _get(data, "ipi_tipo", default="%") or "%"
                    produto.difal = to_number(_get(data, "difal"))
                    produto.imposto_venda = to_number(_get(data, "imposto_venda"))

                    produto.calcular_precos()

            elif filename.endswith(".csv"):
                # Detectar encoding/delimitador e ler linhas
                text = TextIOWrapper(file.stream, encoding="utf-8-sig", newline="")
                try:
                    sniffer = csv.Sniffer()
                    sample = text.read(2048)
                    text.seek(0)
                    dialect = sniffer.sniff(sample)
                except Exception:
                    text.seek(0)
                    dialect = csv.excel

                reader = csv.DictReader(text, dialect=dialect)
                # normalizar headers para lower
                field_map = {h: h.strip().lower() for h in reader.fieldnames or []}

                for row in reader:
                    data = {field_map.get(k, k).lower(): v for k, v in row.items()}

                    sku = (str(data.get("sku") or "").strip().upper())
                    if not sku:
                        continue

                    produto = Produto.query.filter_by(sku=sku).first()
                    if not produto:
                        produto = Produto(sku=sku)
                        db.session.add(produto)
                        criados += 1
                    else:
                        atualizados += 1

                    produto.nome = data.get("nome", produto.nome)
                    produto.preco_fornecedor = to_number(data.get("preco_fornecedor"))
                    produto.desconto_fornecedor = to_number(data.get("desconto_fornecedor"))
                    produto.margem = to_number(data.get("margem"))
                    produto.lucro_alvo = to_number(data.get("lucro_alvo")) or None
                    produto.preco_final = to_number(data.get("preco_final")) or None

                    produto.ipi = to_number(data.get("ipi"))
                    produto.ipi_tipo = data.get("ipi_tipo") or "%"
                    produto.difal = to_number(data.get("difal"))
                    produto.imposto_venda = to_number(data.get("imposto_venda"))

                    produto.calcular_precos()
            else:
                flash("Formato de arquivo não suportado. Use .csv ou .xlsx", "danger")
                return redirect(url_for("main.importar_produtos"))

            db.session.commit()
            flash(f"Importação concluída! {criados} criados, {atualizados} atualizados.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao importar: {e}", "danger")

        return redirect(url_for("main.produtos"))

    return render_template("produtos_importar.html")

@main.route("/produtos/exemplo-csv")
@login_required
def exemplo_csv():
    pasta = os.path.join(os.path.dirname(__file__), "..", "static")
    return send_from_directory(pasta, "exemplo_produtos.csv", as_attachment=True)

# --- Taxas ---
@main.route("/taxas")
@login_required
def taxas():
    taxas = Taxa.query.order_by(Taxa.numero_parcelas).all()
    return render_template("taxas.html", taxas=taxas)

@main.route("/taxa/nova", methods=["GET", "POST"])
@main.route("/taxa/editar/<int:taxa_id>", methods=["GET", "POST"])
@login_required
def gerenciar_taxa(taxa_id=None):
    taxa = Taxa.query.get(taxa_id) if taxa_id else None
    if request.method == "POST":
        if not taxa:
            taxa = Taxa()
            db.session.add(taxa)

        numero_parcelas_raw = request.form.get("numero_parcelas")
        juros_raw = request.form.get("juros")

        taxa.numero_parcelas = int(numero_parcelas_raw or (taxa.numero_parcelas or 1))
        taxa.juros = to_float(juros_raw, default=(taxa.juros or 0))

        db.session.commit()
        flash("Taxa salva com sucesso!", "success")
        return redirect(url_for("main.taxas"))

    return render_template("taxa_form.html", taxa=taxa)

@main.route("/taxa/excluir/<int:taxa_id>")
@login_required
def excluir_taxa(taxa_id):
    taxa = Taxa.query.get_or_404(taxa_id)
    db.session.delete(taxa)
    db.session.commit()
    flash("Taxa excluída com sucesso!", "success")
    return redirect(url_for("main.taxas"))

# --- Parcelamento ---
@main.route("/parcelamento")
@login_required
def parcelamento_index():
    produtos = Produto.query.order_by(Produto.nome).all()
    return render_template("parcelamento_index.html", produtos=produtos)

@main.route("/parcelamento/<int:produto_id>")
@login_required
def parcelamento(produto_id):
    produto = Produto.query.get_or_404(produto_id)
    taxas = Taxa.query.order_by(Taxa.numero_parcelas).all()

    valor_base = produto.preco_final or produto.preco_a_vista or 0.0
    linhas = parc.gerar_linhas_parcelas(valor_base, taxas)

    texto_whats = gerar_mensagem_whatsapp(produto, valor_base, linhas)

    return render_template(
        "parcelamento.html",
        produto=produto,
        resultado=linhas,
        texto_whats=texto_whats
    )

@main.route("/parcelamento/rapido", methods=["GET", "POST"])
@login_required
def parcelamento_rapido():
    resultado = []
    preco_base = None
    texto_whats = ""

    if request.method == "POST":
        preco_base = to_float(request.form.get("preco_base"))
        taxas = Taxa.query.order_by(Taxa.numero_parcelas).all()

        resultado = parc.gerar_linhas_parcelas(preco_base, taxas)
        texto_whats = gerar_mensagem_whatsapp(None, preco_base, resultado)

    return render_template(
        "parcelamento_rapido.html",
        resultado=resultado,
        preco_base=preco_base,
        texto_whats=texto_whats
    )

# --- Configurações ---
@main.route("/configuracoes")
@login_required
def configuracoes():
    configs = Configuracao.query.all()
    return render_template("configuracoes.html", configs=configs)

@main.route("/configuracao/nova", methods=["GET", "POST"])
@main.route("/configuracao/editar/<int:config_id>", methods=["GET", "POST"])
@login_required
def gerenciar_configuracao(config_id=None):
    config = Configuracao.query.get(config_id) if config_id else None
    if request.method == "POST":
        chave = request.form.get("chave")
        valor = request.form.get("valor")

        if not config:
            config = Configuracao(chave=chave, valor=valor)
            db.session.add(config)
        else:
            config.chave = chave
            config.valor = valor

        db.session.commit()
        flash("Configuração salva com sucesso!", "success")
        return redirect(url_for("main.configuracoes"))

    return render_template("configuracao_form.html", config=config)

@main.route("/configuracao/excluir/<int:config_id>")
@login_required
def excluir_configuracao(config_id):
    config = Configuracao.query.get_or_404(config_id)
    db.session.delete(config)
    db.session.commit()
    flash("Configuração excluída com sucesso!", "success")
    return redirect(url_for("main.configuracoes"))

# --- Usuários ---
@main.route("/usuarios")
@login_required
def usuarios():
    users = User.query.all()
    return render_template("usuarios.html", users=users)

@main.route("/usuario/novo", methods=["GET", "POST"])
@main.route("/usuario/editar/<int:user_id>", methods=["GET", "POST"])
@login_required
def gerenciar_usuario(user_id=None):
    user = User.query.get(user_id) if user_id else None
    if request.method == "POST":
        username = request.form.get("username")
        password = request.form.get("password")

        if not user:
            user = User(username=username, password=password)
            db.session.add(user)
        else:
            user.username = username
            if password:
                user.password = password

        db.session.commit()
        flash("Usuário salvo com sucesso!", "success")
        return redirect(url_for("main.usuarios"))

    return render_template("usuario_form.html", user=user)

@main.route("/usuario/excluir/<int:user_id>")
@login_required
def excluir_usuario(user_id):
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    flash("Usuário excluído com sucesso!", "success")
    return redirect(url_for("main.usuarios"))

# --- Health Check ---
@main.route("/health", methods=["GET", "HEAD"])
def health():
    """Rota para UptimeRobot monitorar a aplicação e o banco"""
    try:
        db.session.execute(text("SELECT 1"))
        return {"status": "ok"}, 200
    except Exception as e:
        return {"status": "error", "msg": str(e)}, 500

# --- Importar Relatórios (Clientes e Vendas) ---
@main.route("/importar", methods=["GET", "POST"])
@login_required
def importar():
    """
    Página para importar:
      - Clientes: .xlsx com cabeçalhos do relatório 'lista-de-pessoas'
      - Vendas:   .xlsx com cabeçalhos do relatório 'vendas gerais'
    """
    if request.method == "POST":
        file = request.files.get("file")
        tipo = request.form.get("tipo")

        if not file or file.filename == "":
            flash("Nenhum arquivo selecionado.", "danger")
            return redirect(url_for("main.importar"))

        try:
            if str(file.filename).lower().endswith(".xlsx"):
                if tipo == "clientes":
                    importar_clientes(file)
                    flash("Clientes importados com sucesso!", "success")
                elif tipo == "vendas":
                    importar_vendas(file)
                    flash("Vendas importadas com sucesso!", "success")
                else:
                    flash("Tipo de importação inválido.", "danger")
            else:
                flash("Envie um arquivo .xlsx exportado do sistema.", "danger")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao importar: {e}", "danger")

        return redirect(url_for("main.importar"))

    return render_template("importar.html")

# =========================
# Funções de Importação (openpyxl)
# =========================
def importar_clientes(file_storage):
    """
    Importa clientes de um .xlsx (primeira aba).
    Colunas esperadas (varia conforme relatório, mas cobrimos aliases):
      - Nome Razão Social / Nome
      - Documento (CPF / CNPJ)
      - E-mail, Telefone, Celular
      - Endereço, Número, Complemento, Bairro, Cidade, Estado, CEP
      - RG, RG emissor
      - Profissão, Sexo
      - CR, CR Emissor, SIGMA, SINARM
      - CAC, FILIADO, POLICIAL, BOMBEIRO, MILITAR, IAT, PSICOLOGO
    """
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    headers = _headers_lower(ws)

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = _row_as_dict(headers, row)

        nome = _get(data, "nome razão social", "nome")
        if not nome:
            continue

        doc = str(_get(data, "documento (cpf / cnpj)", "documento") or "").strip()
        if not doc:
            doc = None  # permite NULL no banco, não conflita com UNIQUE

        cliente = Cliente.query.filter_by(documento=doc).first() if doc else None
        if not cliente:
            cliente = Cliente(nome=nome, documento=doc)
            db.session.add(cliente)

        cliente.nome = nome
        cliente.razao_social = _get(data, "razão social", "razao social", default="")
        cliente.sexo = _get(data, "sexo", default="")
        cliente.profissao = _get(data, "profissão", "profissao", default="")
        cliente.rg = _get(data, "rg", default="")
        cliente.rg_emissor = _get(data, "rg emissor", default="")
        cliente.email = _get(data, "e-mail", "email", default="")
        cliente.telefone = _get(data, "telefone", default="")
        cliente.celular = _get(data, "celular", default="")
        cliente.endereco = _get(data, "endereço", "endereco", default="")
        cliente.numero = str(_get(data, "número", "numero", default="") or "")
        cliente.complemento = _get(data, "complemento", default="")
        cliente.bairro = _get(data, "bairro", default="")
        cliente.cidade = _get(data, "cidade", default="")
        cliente.estado = _get(data, "estado", default="")
        cliente.cep = _get(data, "cep", default="")
        cliente.cr = _get(data, "cr", default="")
        cliente.cr_emissor = _get(data, "cr emissor", default="")
        cliente.sigma = _get(data, "sigma", default="")
        cliente.sinarm = _get(data, "sinarm", default="")
        cliente.cac = _as_bool(_get(data, "cac"))
        cliente.filiado = _as_bool(_get(data, "filiado"))
        cliente.policial = _as_bool(_get(data, "policial"))
        cliente.bombeiro = _as_bool(_get(data, "bombeiro"))
        cliente.militar = _as_bool(_get(data, "militar"))
        cliente.iat = _as_bool(_get(data, "iat"))
        cliente.psicologo = _as_bool(_get(data, "psicologo"))

    db.session.commit()

def importar_vendas(file_storage):
    """
    Importa vendas de um .xlsx (primeira aba).
    O relatório costuma vir "em blocos":
      - Linha com informações da venda (tem 'Consumidor')
      - Linhas subsequentes com itens (sem 'Consumidor', mas com 'Produto')
    Este parser cria a venda quando encontra 'Consumidor' e associa itens até a próxima venda.
    """
    wb = load_workbook(file_storage, data_only=True)
    ws = wb.active
    headers = _headers_lower(ws)

    current_venda = None

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = _row_as_dict(headers, row)

        consumidor = _get(data, "consumidor")
        documento = str(_get(data, "documento") or "").strip()
        if documento == "":
            documento = None

        # Se a linha tem 'Consumidor', abre uma nova venda
        if consumidor:
            # garantir cliente
            if documento:
                cliente = Cliente.query.filter_by(documento=documento).first()
                if not cliente:
                    cliente = Cliente(nome=consumidor or "", documento=documento)
                    db.session.add(cliente)
                    db.session.flush()
            else:
                # Reutiliza/Cria cliente genérico sem documento
                cliente = Cliente.query.filter_by(documento=None, nome="Consumidor não identificado").first()
                if not cliente:
                    cliente = Cliente(nome="Consumidor não identificado", documento=None)
                    db.session.add(cliente)
                    db.session.flush()

            current_venda = Venda(
                cliente_id=cliente.id,
                vendedor=_get(data, "vendedor"),
                status=_get(data, "status"),
                status_financeiro=_get(data, "status financeiro"),
                data_abertura=parse_data(_get(data, "abertura")),
                data_fechamento=parse_data(_get(data, "fechamento")),
                data_quitacao=parse_data(_get(data, "quitação", "quitacao")),
                valor_total=to_number(_get(data, "valor total")),
                nf_numero=str(_get(data, "nf - nº", "nf-nº", "nf nº", default="") or ""),
                nf_valor=to_number(_get(data, "nf - valor", "nf valor")),
                teve_devolucao=_as_bool(_get(data, "teve devoluções", "teve devolucoes")),
            )
            db.session.add(current_venda)
            db.session.flush()

            # Se a própria linha já tiver produto, cria item
            produto_nome = _get(data, "produto")
            if produto_nome:
                qtd = int(to_number(_get(data, "itens - qtd", "qtd", default=1)) or 1)
                valor = to_number(_get(data, "valor"))
                item = ItemVenda(
                    venda_id=current_venda.id,
                    produto_nome=produto_nome,
                    categoria=_get(data, "tipo do produto", "categoria", default=""),
                    quantidade=qtd,
                    valor_unitario=valor,
                    valor_total=valor * qtd,
                )
                db.session.add(item)

            continue  # próxima linha

        # Se NÃO tem consumidor, mas tem produto -> é item da venda atual
        if current_venda and _get(data, "produto"):
            qtd = int(to_number(_get(data, "itens - qtd", "qtd", default=1)) or 1)
            valor = to_number(_get(data, "valor"))
            item = ItemVenda(
                venda_id=current_venda.id,
                produto_nome=_get(data, "produto", default=""),
                categoria=_get(data, "tipo do produto", "categoria", default=""),
                quantidade=qtd,
                valor_unitario=valor,
                valor_total=valor * qtd,
            )
            db.session.add(item)

    db.session.commit()

# ---------------------------------------------------
# Funções auxiliares
# ---------------------------------------------------
def parse_brl(s) -> float:
    if s is None:
        return 0.0
    s = str(s).strip()
    if not s:
        return 0.0
    s = s.replace("R$", "").replace("r$", "").strip()
    s = s.replace(".", "")
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0

def parse_pct(s) -> float:
    if s is None:
        return 0.0
    s = str(s).strip().replace("%", "").replace(",", ".")
    if not s:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0

# ---------------------------------------------------
# LISTAR PEDIDOS (com filtros)
# ---------------------------------------------------
@main.route("/pedidos")
@login_required
def listar_pedidos():
    query = PedidoCompra.query.join(PedidoCompra.fornecedor)

    numero = request.args.get("numero", "").strip()
    fornecedor_nome = request.args.get("fornecedor", "").strip()
    data_inicio = request.args.get("data_inicio", "").strip()
    data_fim = request.args.get("data_fim", "").strip()
    valor_min = request.args.get("valor_min", "").strip()
    valor_max = request.args.get("valor_max", "").strip()

    # filtro por número
    if numero:
        query = query.filter(PedidoCompra.numero.ilike(f"%{numero}%"))

    # filtro por fornecedor (nome)
    if fornecedor_nome:
        query = query.filter(Cliente.nome.ilike(f"%{fornecedor_nome}%"))

    # filtro por período
    if data_inicio:
        try:
            di = datetime.strptime(data_inicio, "%Y-%m-%d").date()
            query = query.filter(PedidoCompra.data_pedido >= di)
        except ValueError:
            flash("Data inicial inválida.", "warning")
    if data_fim:
        try:
            df = datetime.strptime(data_fim, "%Y-%m-%d").date()
            query = query.filter(PedidoCompra.data_pedido <= df)
        except ValueError:
            flash("Data final inválida.", "warning")

    # filtro por valores
    if valor_min or valor_max:
        from sqlalchemy import func
        subq = (
            db.session.query(
                ItemPedido.pedido_id,
                func.sum(ItemPedido.quantidade * ItemPedido.valor_unitario).label("subtotal")
            )
            .group_by(ItemPedido.pedido_id)
            .subquery()
        )
        query = query.join(subq, PedidoCompra.id == subq.c.pedido_id)
        if valor_min:
            query = query.filter(subq.c.subtotal >= float(valor_min))
        if valor_max:
            query = query.filter(subq.c.subtotal <= float(valor_max))

    pedidos = query.order_by(PedidoCompra.id.desc()).all()
    return render_template("pedidos/listar.html", pedidos=pedidos)


# ---------------------------------------------------
# NOVO PEDIDO
# ---------------------------------------------------
@main.route("/pedidos/novo", methods=["GET", "POST"])
@login_required
def novo_pedido():
    if request.method == "POST":
        fornecedor_id = request.form.get("fornecedor")
        cond_pagto = request.form.get("cond_pagto")
        modo = request.form.get("modo_desconto")

        perc_armas = parse_pct(request.form.get("percentual_armas"))
        perc_municoes = parse_pct(request.form.get("percentual_municoes"))
        perc_unico = parse_pct(request.form.get("percentual_unico"))

        numero = datetime.now().strftime("%Y%m%d%H%M%S")

        pedido = PedidoCompra(
            numero=numero,
            data_pedido=datetime.now().date(),
            cond_pagto=cond_pagto,
            modo_desconto=modo,
            percentual_armas=perc_armas,
            percentual_municoes=perc_municoes,
            percentual_unico=perc_unico,
            fornecedor_id=int(fornecedor_id) if fornecedor_id else None,
        )
        db.session.add(pedido)
        db.session.flush()

        codigos = request.form.getlist("codigo[]")
        descricoes = request.form.getlist("descricao[]")
        quantidades = request.form.getlist("quantidade[]")
        valores = request.form.getlist("valor_unitario[]")

        for codigo, desc, qtd, val in zip(codigos, descricoes, quantidades, valores):
            desc = (desc or "").strip()
            if not desc:
                continue
            q = int(qtd or 0)
            v = parse_brl(val)
            if q <= 0 or v <= 0:
                continue
            item = ItemPedido(
                pedido_id=pedido.id,
                codigo=codigo,
                descricao=desc,
                quantidade=q,
                valor_unitario=v,
            )
            db.session.add(item)

        db.session.commit()
        flash("Pedido criado com sucesso!", "success")
        return redirect(url_for("main.listar_pedidos"))

    fornecedores = [c for c in Cliente.query.all() if getattr(c, "documento", None) and "/" in c.documento]
    return render_template("pedidos/novo.html", fornecedores=fornecedores)


# ---------------------------------------------------
# EDITAR PEDIDO
# ---------------------------------------------------
@main.route("/pedidos/<int:id>/editar", methods=["GET", "POST"])
@login_required
def editar_pedido(id):
    pedido = PedidoCompra.query.get_or_404(id)

    if request.method == "POST":
        pedido.cond_pagto = request.form.get("cond_pagto")
        pedido.modo_desconto = request.form.get("modo_desconto")
        pedido.percentual_armas = float(request.form.get("percentual_armas") or 0)
        pedido.percentual_municoes = float(request.form.get("percentual_municoes") or 0)
        pedido.percentual_unico = float(request.form.get("percentual_unico") or 0)

        for it in list(pedido.itens):
            db.session.delete(it)

        codigos = request.form.getlist("codigo[]")
        descricoes = request.form.getlist("descricao[]")
        quantidades = request.form.getlist("quantidade[]")
        valores = request.form.getlist("valor_unitario[]")

        for codigo, desc, qtd, val in zip(codigos, descricoes, quantidades, valores):
            if not desc.strip():
                continue
            q = int(qtd or 0)
            v = parse_brl(val)
            if q <= 0 or v <= 0:
                continue
            item = ItemPedido(
                pedido_id=pedido.id,
                codigo=codigo,
                descricao=desc.strip(),
                quantidade=q,
                valor_unitario=v,
            )
            db.session.add(item)

        db.session.commit()
        flash("Pedido atualizado com sucesso!", "success")
        return redirect(url_for("main.listar_pedidos"))

    fornecedores = Cliente.query.all()
    return render_template("pedidos/novo.html", pedido=pedido, fornecedores=fornecedores)

# ---------------------------------------------------
# PDF DO PEDIDO
# ---------------------------------------------------
@main.route("/pedidos/<int:id>/pdf")
@login_required
def pedido_pdf(id):
    pedido = PedidoCompra.query.get_or_404(id)

    # Monta lista de itens no formato esperado pelo gerar_pedido_m4
    itens = []
    for i in pedido.itens:
        itens.append((i.codigo, i.descricao, i.quantidade, i.valor_unitario))

    # Nome do arquivo e diretório de saída
    filename = f"pedido_{pedido.numero}.pdf"
    folder = os.path.join(current_app.root_path, "static", "pdf")
    filepath = os.path.join(folder, filename)
    os.makedirs(folder, exist_ok=True)

    # Dados do fornecedor (Cliente vinculado ao PedidoCompra)
    fornecedor = pedido.fornecedor
    fornecedor_nome = fornecedor.nome or ""
    fornecedor_cnpj = fornecedor.documento or ""
    fornecedor_endereco = f"{(fornecedor.endereco or '')}, {(fornecedor.numero or '')} - {(fornecedor.cidade or '')}/{(fornecedor.estado or '')}"
    fornecedor_cr = f"CR {fornecedor.cr or ''}"

    # Gera PDF com os dados reais
    gerar_pedido_m4(
        itens=itens,
        cond_pagto=pedido.cond_pagto,
        perc_armas=pedido.percentual_armas,
        perc_municoes=pedido.percentual_municoes,
        perc_unico=pedido.percentual_unico,
        modo=pedido.modo_desconto,
        numero_pedido=pedido.numero,
        data_pedido=pedido.data_pedido.strftime("%d/%m/%Y") if pedido.data_pedido else None,
        fornecedor_nome=fornecedor_nome,
        fornecedor_cnpj=fornecedor_cnpj,
        fornecedor_endereco=fornecedor_endereco,
        fornecedor_cr=fornecedor_cr,
    )

    # Move o PDF gerado para a pasta correta
    if os.path.exists("pedido_m4.pdf"):
        os.replace("pedido_m4.pdf", filepath)

    return send_file(filepath, as_attachment=False)

# ---------------------------------------------------
# EXCLUIR PEDIDO
# ---------------------------------------------------
@main.route("/pedidos/<int:id>/excluir", methods=["POST", "GET"])
@login_required
def excluir_pedido(id):
    pedido = PedidoCompra.query.get_or_404(id)
    db.session.delete(pedido)
    db.session.commit()
    flash("Pedido excluído com sucesso!", "success")
    return redirect(url_for("main.listar_pedidos"))

# =========================
# API WhatsApp (Produto)
# =========================
@main.route("/api/produto/<int:produto_id>/whatsapp")
@login_required
def api_produto_whatsapp(produto_id):
    produto = Produto.query.get_or_404(produto_id)
    taxas = Taxa.query.order_by(Taxa.numero_parcelas).all()

    valor_base = produto.preco_final or produto.preco_a_vista or 0.0
    linhas = parc.gerar_linhas_parcelas(valor_base, taxas)

    texto_whats = gerar_mensagem_whatsapp(produto, valor_base, linhas)

    return jsonify({"texto_completo": texto_whats})
